"""Spreadsheet providers (CSV / XLSX) for the Constellation pipeline.

Both convert tabular files into Stage-1 Block objects, reusing the existing
table block lane (table_data + Markdown text):

- CsvProvider  -> a single table block (encoding + delimiter sniffing,
  zero third-party deps).
- XlsxProvider -> per worksheet a heading text block (the sheet name)
  followed by a table block, with merged-cell fill-down, formula value
  resolution and Excel number-format rendering.

Large-table policy (deliberately positioned for small/medium tables):
the renderer never emits an unusable mega-table. When a table exceeds the
render budget (or hits the read cap), the Markdown becomes a head+tail
*sample* with an explicit in-band warning and a column overview, while
``table_data.rows`` still carries the full rows that were read. This keeps
output consumable by an LLM and stops silent data loss.

Why two classes instead of one sniffing magic bytes: the API layer's
extract_from_bytes never sees the filename, so format selection has to
happen via the provider registry keyed by source_format.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from app.core.exceptions import ProviderError
from infrastructure.models import Block

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised only without the optional dep
    openpyxl = None  # type: ignore[assignment]


# Read cap: bounds memory for one table. Rows beyond this are not loaded.
_DEFAULT_MAX_ROWS = 5000
_DEFAULT_MAX_COLS = 256
# Render budget: how many rows the Markdown view may contain before it is
# downgraded to a head+tail sample (keeps LLM token cost bounded).
_DEFAULT_MAX_RENDER_ROWS = 200
_DEFAULT_SAMPLE_TAIL = 20

# Delimiter sniffing reads a generous head so multi-line records or sparse
# leading rows don't fool the detector.
_CSV_SNIFF_SAMPLE = 65536

_CSV_ENCODINGS = (
    "utf-8-sig",
    "utf-8",
    "utf-16",
    "gb18030",
    "gbk",
    "big5",
    "latin-1",
)

_CURRENCY_SYMBOLS = "¥$€£￥₩"
_DECIMALS_RE = re.compile(r"\.([0#]+)")


def _stringify(value: object) -> str:
    """Coerce a spreadsheet cell value into a clean display string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, datetime):
        has_time = bool(value.hour or value.minute or value.second or value.microsecond)
        return value.isoformat(sep=" ") if has_time else value.date().isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)


def _decimals_from_format(number_format: str) -> int:
    """Count the decimal-place digits declared in an Excel number format."""
    match = _DECIMALS_RE.search(number_format)
    return len(match.group(1)) if match else 0


def _currency_prefix(number_format: str) -> str:
    """Return the first bare currency symbol found in a number format, if any."""
    for ch in number_format:
        if ch in _CURRENCY_SYMBOLS:
            return ch
    return ""


def _apply_number_format(value: float, number_format: str) -> Optional[str]:
    """Render a numeric cell per a pragmatic subset of Excel number formats.

    Handles percent, thousands separators, a leading currency symbol and a
    fixed number of decimals. Returns ``None`` when the format carries no such
    signal, so the caller falls back to the plain :func:`_stringify` form.
    """
    if "%" in number_format:
        decimals = _decimals_from_format(number_format)
        return f"{value * 100:.{decimals}f}%"

    prefix = _currency_prefix(number_format)
    has_thousands = "#,#" in number_format or ",0" in number_format
    decimals = _decimals_from_format(number_format)

    if prefix or has_thousands:
        body = f"{value:,.{decimals}f}" if has_thousands else f"{value:.{decimals}f}"
        return f"{prefix}{body}"
    if decimals > 0 and ("0" in number_format or "#" in number_format):
        return f"{value:.{decimals}f}"
    return None


def _format_cell_value(value: object, number_format: Optional[str]) -> str:
    """Stringify a cell, applying Excel number formatting to plain numerics."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        fmt = number_format or "General"
        if fmt != "General":
            formatted = _apply_number_format(float(value), fmt)
            if formatted is not None:
                return formatted
    return _stringify(value)


def _trim_empty_columns(rows: List[List[str]]) -> List[List[str]]:
    """Drop trailing all-empty columns; keep interior gaps (may be meaningful)."""
    if not rows:
        return rows
    width = max((len(r) for r in rows), default=0)
    last_nonempty = -1
    for c in range(width):
        if any(c < len(r) and r[c] for r in rows):
            last_nonempty = c
    if last_nonempty < 0:
        return []
    return [r[: last_nonempty + 1] for r in rows]


def _rows_to_markdown(rows: List[List[str]]) -> str:
    """Render string rows as a GFM table via the pipeline's single entry point.

    Delegates to :meth:`Block.render_markdown_table` so spreadsheet output shares
    the same robustness as DOCX/PDF tables: max-width alignment, no truncation of
    over-wide rows, in-cell newline -> ``<br>`` (kept losslessly instead of being
    squashed to a space), and pipe escaping.
    """
    return Block.render_markdown_table(rows)


def _looks_numeric(text: str) -> bool:
    """Heuristic: does a cell string represent a number (after format strip)?"""
    t = text.strip().replace(",", "")
    for ch in _CURRENCY_SYMBOLS:
        t = t.replace(ch, "")
    t = t.rstrip("%")
    if not t:
        return False
    try:
        float(t)
        return True
    except ValueError:
        return False


def _column_overview(header: List[str], data: List[List[str]]) -> List[Dict[str, object]]:
    """Build a lightweight per-column profile (name / kind / non-empty count)."""
    width = max([len(header)] + [len(r) for r in data], default=0)
    columns: List[Dict[str, object]] = []
    for c in range(width):
        name = header[c] if c < len(header) and header[c] else f"col{c + 1}"
        nonempty = 0
        numeric = 0
        for row in data:
            if c < len(row) and row[c]:
                nonempty += 1
                if _looks_numeric(row[c]):
                    numeric += 1
        kind = "numeric" if nonempty and numeric >= nonempty * 0.8 else "text"
        columns.append({"name": name, "kind": kind, "nonempty": nonempty})
    return columns


def _overview_line(columns: List[Dict[str, object]]) -> str:
    parts = [f"{c['name']}({c['kind']})" for c in columns[:20]]
    suffix = "" if len(columns) <= 20 else f" …共 {len(columns)} 列"
    return "> 列概览：" + " | ".join(parts) + suffix


def _warning_line(
    total_rows: int,
    ncols: int,
    head_shown: int,
    tail_shown: int,
    read_truncated: bool,
    read_rows: int,
) -> str:
    approx = "约 " if read_truncated else ""
    tail_part = f" + 后 {tail_shown}" if tail_shown else ""
    base = (
        f"> ⚠️ 大表采样：共 {approx}{total_rows} 行 × {ncols} 列，"
        f"本视图仅展示前 {head_shown}{tail_part} 行（采样供 LLM 概览）。"
    )
    if read_truncated:
        base += f" 已超读取上限，仅载入前 {read_rows} 行，其余未读取。"
    else:
        base += " 完整数据见 table_data.rows。"
    return base


def _assemble_table_block(
    block_id: int,
    rows: List[List[str]],
    total_rows: int,
    read_truncated: bool,
    source_meta: Dict[str, object],
    max_render_rows: int,
    sample_tail: int,
    caption: Optional[str] = None,
) -> Block:
    """Build a table Block, downgrading oversized tables to a sampled view."""
    ncols = max((len(r) for r in rows), default=0)
    header = rows[0] if rows else []
    data = rows[1:]
    columns = _column_overview(header, data)
    needs_sampling = read_truncated or len(rows) > max_render_rows

    if not needs_sampling:
        text = _rows_to_markdown(rows)
    else:
        head_n = max(max_render_rows - sample_tail, 1)
        head_data = data[:head_n]
        if sample_tail > 0 and len(data) > head_n + sample_tail:
            tail_data = data[-sample_tail:]
        else:
            tail_data = []
        shown = [header] + head_data + tail_data
        warning = _warning_line(
            total_rows, ncols, len(head_data), len(tail_data), read_truncated, len(rows)
        )
        text = f"{warning}\n{_overview_line(columns)}\n\n{_rows_to_markdown(shown)}"

    metadata = dict(source_meta)
    metadata.update(
        {
            "row_count": len(rows),
            "total_rows": total_rows,
            "truncated": read_truncated,
            "sampled": needs_sampling,
            "columns": columns,
        }
    )
    return Block(
        id=block_id,
        type="table",
        text=text,
        table_data={"rows": rows},
        caption=caption,
        metadata=metadata,
    )


class CsvProvider:
    """Convert a CSV file into a single Stage-1 table block."""

    def __init__(
        self,
        max_rows: int = _DEFAULT_MAX_ROWS,
        max_cols: int = _DEFAULT_MAX_COLS,
        max_render_rows: int = _DEFAULT_MAX_RENDER_ROWS,
        sample_tail: int = _DEFAULT_SAMPLE_TAIL,
    ) -> None:
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.max_render_rows = max_render_rows
        self.sample_tail = sample_tail

    def extract(self, file_path: str) -> List[Block]:
        path = Path(file_path)
        if path.suffix.lower() != ".csv":
            raise ProviderError("Only .csv files are supported")
        return self.extract_from_bytes(path.read_bytes())

    def extract_from_bytes(self, file_bytes: bytes) -> List[Block]:
        text = self._decode(file_bytes)
        all_rows = self._parse_rows(text)
        if not all_rows:
            return []

        total_rows = len(all_rows)
        read_truncated = total_rows > self.max_rows
        kept = all_rows[: self.max_rows]
        kept = [[_stringify(c) for c in row[: self.max_cols]] for row in kept]
        kept = _trim_empty_columns(kept)
        if not kept:
            return []

        return [
            _assemble_table_block(
                0,
                kept,
                total_rows,
                read_truncated,
                {"source": "csv"},
                self.max_render_rows,
                self.sample_tail,
            )
        ]

    def _decode(self, file_bytes: bytes) -> str:
        for encoding in _CSV_ENCODINGS:
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise ProviderError("Unable to decode CSV; use UTF-8, UTF-16, or GB encodings")

    def _parse_rows(self, text: str) -> List[List[str]]:
        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        delimiter = ","
        try:
            dialect = csv.Sniffer().sniff(
                normalized[:_CSV_SNIFF_SAMPLE], delimiters=",;\t|"
            )
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
        reader = csv.reader(io.StringIO(normalized), delimiter=delimiter)
        return [row for row in reader if any(cell.strip() for cell in row)]


class XlsxProvider:
    """Convert an XLSX workbook into per-sheet heading + table blocks."""

    def __init__(
        self,
        max_rows: int = _DEFAULT_MAX_ROWS,
        max_cols: int = _DEFAULT_MAX_COLS,
        max_render_rows: int = _DEFAULT_MAX_RENDER_ROWS,
        sample_tail: int = _DEFAULT_SAMPLE_TAIL,
    ) -> None:
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.max_render_rows = max_render_rows
        self.sample_tail = sample_tail

    def extract(self, file_path: str) -> List[Block]:
        path = Path(file_path)
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise ProviderError("Only .xlsx files are supported")
        return self.extract_from_bytes(path.read_bytes())

    def extract_from_bytes(self, file_bytes: bytes) -> List[Block]:
        if openpyxl is None:
            raise ProviderError("请安装 openpyxl: pip install openpyxl")
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(file_bytes), data_only=True, read_only=False
            )
        except Exception as exc:  # noqa: BLE001 - normalize into ProviderError
            raise ProviderError(f"无法解析 XLSX 文件: {exc}") from exc

        blocks: List[Block] = []
        block_id = 0
        try:
            for worksheet in workbook.worksheets:
                sheet_name = str(worksheet.title)
                blocks.append(
                    Block(
                        id=block_id,
                        type="text",
                        text=sheet_name,
                        is_heading_style=True,
                        heading_level=1,
                        metadata={"source": "xlsx", "sheet": sheet_name},
                    )
                )
                block_id += 1

                rows, read_truncated, total_rows = self._sheet_to_rows(worksheet)
                if not rows:
                    continue

                blocks.append(
                    _assemble_table_block(
                        block_id,
                        rows,
                        total_rows,
                        read_truncated,
                        {"source": "xlsx", "sheet": sheet_name},
                        self.max_render_rows,
                        self.sample_tail,
                        caption=sheet_name,
                    )
                )
                block_id += 1
        finally:
            try:
                workbook.close()
            except Exception:  # noqa: BLE001 - best-effort cleanup
                pass

        return blocks

    def _sheet_to_rows(self, worksheet) -> Tuple[List[List[str]], bool, int]:
        try:
            merged_ranges = list(worksheet.merged_cells.ranges)
        except Exception:  # noqa: BLE001 - some workbooks expose no merge info
            merged_ranges = []

        values: List[List[object]] = []
        formats: List[List[Optional[str]]] = []
        hit_cap = False
        for row_index, row in enumerate(worksheet.iter_rows()):
            if row_index >= self.max_rows:
                hit_cap = True
                break
            row_values: List[object] = []
            row_formats: List[Optional[str]] = []
            for cell in row[: self.max_cols]:
                row_values.append(cell.value)
                row_formats.append(cell.number_format)
            values.append(row_values)
            formats.append(row_formats)
        if not values:
            return [], False, 0

        self._apply_merged_cells(merged_ranges, values, formats)
        str_rows = [
            [_format_cell_value(values[r][c], formats[r][c]) for c in range(len(values[r]))]
            for r in range(len(values))
        ]
        str_rows = [row for row in str_rows if any(cell for cell in row)]
        str_rows = _trim_empty_columns(str_rows)

        if hit_cap:
            # max_row is an upper-bound estimate (may include trailing blanks).
            total_rows = max(worksheet.max_row or len(str_rows), len(str_rows) + 1)
        else:
            total_rows = len(str_rows)
        return str_rows, hit_cap, total_rows

    def _apply_merged_cells(
        self,
        merged_ranges,
        values: List[List[object]],
        formats: List[List[Optional[str]]],
    ) -> None:
        """Propagate each merged region's top-left value AND format.

        openpyxl only stores a merged value in the anchor cell; the rest read
        as None. Without fill-down the Markdown table would be misaligned. The
        anchor's number format is copied too, so a merged numeric/percent cell
        renders consistently across the region.
        """
        for rng in merged_ranges:
            min_r, min_c = rng.min_row - 1, rng.min_col - 1
            max_r, max_c = rng.max_row - 1, rng.max_col - 1
            if min_r < 0 or min_c < 0 or min_r >= len(values):
                continue
            if min_c >= len(values[min_r]):
                continue
            anchor_value = values[min_r][min_c]
            anchor_format = formats[min_r][min_c]
            for rr in range(min_r, min(max_r + 1, len(values))):
                for cc in range(min_c, min(max_c + 1, len(values[rr]))):
                    values[rr][cc] = anchor_value
                    formats[rr][cc] = anchor_format
